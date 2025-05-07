""" 
pip install openpyxl
"""
import openpyxl
import os
import datetime

def writeExcelByOpenpyxl():
    """ 通过openpyxl写入excel """
    workbook = openpyxl.Workbook() # 创建一个workbook对象
    del workbook['Sheet'] # 删除默认自带的sheet 等价于：workbook.remove(workbook['Sheet']) 
    sheet = workbook.create_sheet("student") # 创建一个sheet工作簿对象
    sheet.append(['姓名','年龄','成绩']) # 添加表头
    nameList=['张三','李四','王五']
    ageList=[18,19,20]
    scoreList=[i for i in range(3)]
    zipObj= zip(nameList,ageList,scoreList) # 将打包成元组并返回一个可迭代的 zip 对象 
    for name,age,score in zipObj:
        sheet.append([name,age,score]) # 迭代添加每个列表，一个列表对应工作簿的一行

    # 创建输出目录
    output_path = ".output"
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    filepath = os.path.join(output_path,  f"学生信息-{(datetime.datetime.now().strftime('%Y%m%d'))}.xlsx")  # 拼接完整的文件保存路径
    workbook.save(filepath)

def readExcelByOpenpyxl():
    """ 通过openpyxl读取excel """
    workbook = openpyxl.load_workbook(filename=r"D:\workspace\python-projects\my-pythonapp\.output\学生信息-20250504.xlsx") #raw string，不处理转义字符
    sheet = workbook['student'] # 获取工作簿对象
    rowList=[] # 存储行数据,最后是一个二维列表
    for row in sheet.rows: # 遍历工作簿的每一行
        # rowList.append(row)
        # print(row,type(row)) # (<Cell 'student'.A1>, <Cell 'student'.B1>, <Cell 'student'.C1>) <class 'tuple'>
        rowData = [cell.value for cell in row] # 获取行数据: 每个单元格数据组成的列表 ['张三', 18, 0]
        print(rowData)
        rowList.append(rowData)
    for row in rowList:
        print(row)


